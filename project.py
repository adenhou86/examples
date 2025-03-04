import pandas as pd
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.cell.cell import Hyperlink

# Assuming the DataFrame is named df
# Example DataFrame creation (replace this with your actual DataFrame)
data = {
    'doc_id': [41824151, 41824217, 41824298, 41824311, 41824363, 41824625, 41824633, 41824654, 41824656, 41824659],
    'question': ['What is the client legal name?', 'What is the client legal name?', 'What is the client legal name?', 'What is the client legal name?', 'What is the client legal name?', 
                  "Could you state 'Yes' or 'No' to indicate if t...", "Could you state 'Yes' or 'No' to indicate if t...", "Could you state 'Yes' or 'No' to indicate if t...", 
                  "Could you state 'Yes' or 'No' to indicate if t...", "Could you state 'Yes' or 'No' to indicate if t..."],
    'answer': ['NOT_FOUND', 'Eze Castle Software LLC', 'MELLON ALPHAEQUITY UK FUND, LTD.', 'NOT_FOUND', 'OVT Fund LP', 
               'Yes, the client\'s affiliates are covered by th...', 'NOT_FOUND', 'Yes', 'No', 'NOT_FOUND'],
    'topic': ['Client Legal Name', 'Client Legal Name', 'Client Legal Name', 'Client Legal Name', 'Client Legal Name', 
              'Clients Affiliates Covered', 'Clients Affiliates Covered', 'Clients Affiliates Covered', 'Clients Affiliates Covered', 'Clients Affiliates Covered']
}

df = pd.DataFrame(data)

# Define a border style
border_style = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Define a center alignment style
center_alignment = Alignment(horizontal='center')

# Create a Pandas Excel writer using XlsxWriter as the engine.
with pd.ExcelWriter('output.xlsx', engine='openpyxl') as writer:
    # Group by the 'topic' column and write each group to a separate sheet
    for topic, group in df.groupby('topic'):
        # Limit sheet name to the first 20 characters
        sheet_name = topic[:20]
        group.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Access the worksheet
        worksheet = writer.sheets[sheet_name]
        
        # Iterate through columns and adjust their widths based on content
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)  # Get the column letter (A, B, C, etc.)
            
            # Find the maximum length of content in the column
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set the column width to fit the content (add a little extra space)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply borders to all cells in the table
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border_style
        
        # Center align the column headers
        for cell in worksheet[1]:
            cell.alignment = center_alignment
        
        # Add hyperlinks to the doc_id column (Column A)
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=1):
            for cell in row:
                doc_id = cell.value
                if doc_id:  # Ensure the cell is not empty
                    # Create a hyperlink (replace '#' with the actual URL if needed)
                    cell.hyperlink = f"https://example.com/document/{doc_id}"
                    cell.style = "Hyperlink"  # Apply the hyperlink style

# The file is automatically saved and closed when exiting the 'with' block
